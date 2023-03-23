import calculos
import selenium
import time
import os
import openpyxl
import openpyxl.styles
from openpyxl.utils.cell import get_column_letter
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select


url = "https://marangatu.set.gov.py/eset/login"
descargas = r'C:\Users\Juani\Documents\Downloads'

usuario = 1341595
contrasena = "Austria2022"

desde = "01/01/2022"
hasta = "30/04/2022"

opciones = ['COMPRAS', 'VENTAS', 'INGRESOS', 'EGRESOS']
opcionelegida = 'COMPRAS'

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get(url)

elem = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.NAME, "usuario")))
elem.send_keys(usuario)
elem.send_keys(Keys.TAB)

elem2 = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.NAME, "clave")))
elem2.send_keys(contrasena)
elem2.send_keys(Keys.TAB)

time.sleep(1)
elem3 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "btn-primary")))
elem3.click()

time.sleep(1)
elem4 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div/div/div/div[2]/div/div[1]/div/div/div/div/section/div/div/div[2]/div/div/div[7]/div/div[1]")))
elem4.click()

time.sleep(1)
elem5 = WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div/div/div/div[2]/div/div[1]/div/div/div/div/section/div/div/div[2]/div/div/div[5]/div/div")))
elem5.click()

fechas = calculos.calcularfechas(desde, hasta)

parent_window = driver.current_window_handle
all_windows = driver.window_handles
driver.switch_to.window(all_windows[1])

time.sleep(1)

elem6 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div/div[2]/div/div/div[2]/div/div/section[1]/div/div/form/div/div/div[2]/div[1]/div/div/input")))

elem7 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div/div[2]/div/div/div[2]/div/div/section[1]/div/div/form/div/div/div[2]/div[2]/div/div/input")))

elem8 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div/div[2]/div/div/div[2]/div/div/section[1]/div/div/form/div/div/div[1]/div[2]/div/select")))
select = Select(elem8)
select.select_by_index(2)

elem9 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'btn-primary.btn.ng-scope')))

i = 0

for fechainicio, fechafinal in fechas:
    print(fechainicio, fechafinal)
    elem6.clear()  # clear the input field before sending the date
    elem6.send_keys(fechainicio)
    print(fechainicio)
    elem7.clear()  # clear the input field before sending the date
    elem7.send_keys(fechafinal)
    print(fechafinal)
    elem9 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, 'btn-primary.btn.ng-scope')))
    elem9.click()
    print("busqueda")
    time.sleep(1.5)
    elem10 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "btn.btn-labeled.btn-secondary.mb-2")))
    elem10.click()
    print("descarga")
    time.sleep(1.5)
    # take the latest downloaded file
    files = os.listdir(descargas)
    xlsx_files = [f for f in files if f.endswith('.xlsx')]
    xlsx_files.sort(key=lambda f: os.path.getmtime(os.path.join(descargas, f)))
    latest_file = xlsx_files[-1]
    print(latest_file)

    if(i == 0):
        primerarchivo = latest_file
    else:
        # load both workbooks
        primerarchivo_wb = openpyxl.load_workbook(os.path.join(descargas,primerarchivo))
        current_file_wb = openpyxl.load_workbook(os.path.join(descargas,latest_file))

        # get the first sheet of each workbook
        primerarchivo_sheet = primerarchivo_wb.active
        current_file_sheet = current_file_wb.active

        # append the data from the current file to the first sheet of primerarchivo
        for row in current_file_sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            primerarchivo_sheet.append(values)

        # save the updated workbook
        primerarchivo_wb.save(os.path.join(descargas,primerarchivo))

    i = i + 1


    
       

    
    




    





